"""Load documents (PDFs, images, CSVs, Excel) into Claude API content blocks."""

import base64
import csv
import io
import os
from pathlib import Path

# Supported file extensions
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".tif"}
PDF_EXTENSIONS = {".pdf"}
CSV_EXTENSIONS = {".csv", ".tsv"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}

MIME_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/png",  # convert bmp to png
    ".tiff": "image/png",
    ".tif": "image/png",
}


def load_document(file_path: str) -> list[dict]:
    """Convert a file into a list of Claude API content blocks.

    Returns a list of dicts suitable for use in Claude message content arrays.
    Images become base64-encoded image blocks, text becomes text blocks.
    """
    path = Path(file_path)
    if not path.exists():
        return [{"type": "text", "text": f"[File not found: {file_path}]"}]

    ext = path.suffix.lower()

    if ext in PDF_EXTENSIONS:
        return _load_pdf(path)
    elif ext in IMAGE_EXTENSIONS:
        return _load_image(path)
    elif ext in CSV_EXTENSIONS:
        return _load_csv(path, delimiter="\t" if ext == ".tsv" else ",")
    elif ext in EXCEL_EXTENSIONS:
        return _load_excel(path)
    else:
        # Try to read as plain text
        try:
            text = path.read_text(encoding="utf-8")
            return [{"type": "text", "text": f"[Contents of {path.name}]\n{text}"}]
        except (UnicodeDecodeError, OSError):
            return [{"type": "text", "text": f"[Unsupported file format: {ext}]"}]


def load_documents(file_paths: list[str]) -> list[dict]:
    """Load multiple documents into a flat list of content blocks."""
    blocks = []
    for fp in file_paths:
        blocks.extend(load_document(fp))
    return blocks


def _load_pdf(path: Path) -> list[dict]:
    """Render each PDF page as a PNG image for Claude Vision."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return [{"type": "text", "text": "[PyMuPDF not installed. Cannot process PDF.]"}]

    blocks = []
    doc = fitz.open(str(path))

    for page_num in range(len(doc)):
        page = doc[page_num]
        # Render at 2x resolution for better OCR
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        img_bytes = pix.tobytes("png")
        b64 = base64.standard_b64encode(img_bytes).decode("utf-8")

        blocks.append({
            "type": "text",
            "text": f"[Page {page_num + 1} of {path.name}]",
        })
        blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": b64,
            },
        })

    doc.close()
    return blocks


def _load_image(path: Path) -> list[dict]:
    """Load an image file as a base64 content block."""
    ext = path.suffix.lower()
    media_type = MIME_TYPES.get(ext, "image/png")

    img_bytes = path.read_bytes()

    # For formats Claude doesn't support natively, we'd need conversion
    # but for common formats (png, jpg, gif, webp) just encode directly
    b64 = base64.standard_b64encode(img_bytes).decode("utf-8")

    return [
        {"type": "text", "text": f"[Image: {path.name}]"},
        {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_type,
                "data": b64,
            },
        },
    ]


def _load_csv(path: Path, delimiter: str = ",") -> list[dict]:
    """Read a CSV/TSV file and format as a markdown table."""
    text = path.read_text(encoding="utf-8")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)

    if not rows:
        return [{"type": "text", "text": f"[Empty CSV: {path.name}]"}]

    # Build markdown table
    header = rows[0]
    md_lines = [
        f"[Spreadsheet: {path.name}]",
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in rows[1:]:
        # Pad row to match header length
        padded = row + [""] * (len(header) - len(row))
        md_lines.append("| " + " | ".join(padded[:len(header)]) + " |")

    return [{"type": "text", "text": "\n".join(md_lines)}]


def _load_excel(path: Path) -> list[dict]:
    """Read an Excel file and format as markdown tables (one per sheet)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [{"type": "text", "text": "[openpyxl not installed. Cannot process Excel.]"}]

    wb = load_workbook(str(path), read_only=True, data_only=True)
    blocks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        header = [str(c) if c is not None else "" for c in rows[0]]
        md_lines = [
            f"[Sheet: {sheet_name} from {path.name}]",
            "| " + " | ".join(header) + " |",
            "| " + " | ".join("---" for _ in header) + " |",
        ]
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            padded = cells + [""] * (len(header) - len(cells))
            md_lines.append("| " + " | ".join(padded[:len(header)]) + " |")

        blocks.append({"type": "text", "text": "\n".join(md_lines)})

    wb.close()
    return blocks
